"""
Export Manager - Multi-format data export for network scan results

This module provides comprehensive export capabilities for NetworkMapper,
allowing scan results to be shared in various formats suitable for different
audiences and use cases. It handles data transformation, formatting, and
file generation for professional-quality outputs.

Supported formats:
- PDF: Executive reports with charts and summaries
- Excel: Detailed workbooks with multiple sheets
- CSV: Simple data interchange format
- JSON: Complete data export for integration

Key Features:
- Professional formatting with consistent styling
- Automatic summarization and statistics
- Vulnerability data integration
- Change tracking visualization
- Device categorization by type

Design Philosophy:
- One-click export with sensible defaults
- Rich formatting that's print-ready
- Scalable to handle large networks
- Preserves all relevant metadata
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.platypus.tableofcontents import TableOfContents

logger = logging.getLogger(__name__)


class ExportManager:
    """
    Manages exports to various formats (PDF, Excel, CSV, JSON).

    This class orchestrates the export process, handling format-specific
    requirements while maintaining consistency across all export types.
    Each format serves different needs:
    - PDF: Management reports and documentation
    - Excel: Detailed analysis and inventory
    - CSV: Data import/export for other tools
    - JSON: Full data preservation and APIs
    """

    def __init__(self, output_path: Path = None):
        """
        Initialize export manager with output directory.

        Creates a dedicated exports directory to organize output files.
        This separation keeps exports distinct from scan results and
        other application data.

        Args:
            output_path: Base path for exports (default: current directory)
        """
        self.output_path = output_path or Path(".")
        self.export_path = self.output_path / "exports"
        self.export_path.mkdir(exist_ok=True)

    def export_to_pdf(
        self, devices: List[Dict], changes: Optional[Dict] = None, filename: Optional[str] = None
    ) -> Path:
        """
        Export network scan data to professional PDF report.

        Generates a comprehensive PDF report suitable for management and
        documentation purposes. The report includes:
        - Executive summary with key metrics
        - Device type distribution charts
        - Network change analysis (if provided)
        - Critical infrastructure highlights
        - Detailed device inventory (first 50)

        The PDF uses professional styling with color coding for different
        sections and importance levels. Tables are formatted for clarity
        and reports are designed to be print-friendly.

        Args:
            devices: List of device dictionaries containing scan results
            changes: Optional change tracking data with new/missing/changed devices
            filename: Optional custom filename (auto-generated if not provided)

        Returns:
            Path to generated PDF file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filename or f"network_report_{timestamp}.pdf"
        filepath = self.export_path / filename

        # Create PDF document
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=letter,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            rightMargin=0.75 * inch,
        )

        # Container for elements
        elements = []
        styles = getSampleStyleSheet()

        # Custom styles for professional appearance
        # Blue color scheme maintains consistency with web interface
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#2563eb"),  # Blue for main title
            spaceAfter=30,
            alignment=1,  # Center alignment for title page
        )

        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#1e40af"),
            spaceAfter=12,
        )

        # Title page
        elements.append(Paragraph("Network Scan Report", title_style))
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(
            Paragraph(
                f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles["Normal"]
            )
        )
        elements.append(Spacer(1, 0.5 * inch))

        # Executive Summary
        elements.append(Paragraph("Executive Summary", heading_style))

        # Calculate summary statistics for executive overview
        # These metrics provide quick insights into network composition
        device_types = {}  # Count devices by type
        critical_count = 0  # Track critical infrastructure
        for device in devices:
            dtype = device.get("type", "unknown")
            device_types[dtype] = device_types.get(dtype, 0) + 1
            if device.get("critical", False):
                critical_count += 1

        summary_data = [
            ["Metric", "Value"],
            ["Total Devices", str(len(devices))],
            ["Critical Devices", str(critical_count)],
            ["Device Types", str(len(device_types))],
            ["Scan Date", datetime.now().strftime("%Y-%m-%d")],
        ]

        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f3f4f6")),
                    ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#e5e7eb")),
                ]
            )
        )

        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Device Type Distribution
        elements.append(Paragraph("Device Type Distribution", heading_style))

        type_data = [["Device Type", "Count", "Percentage"]]
        for dtype, count in sorted(device_types.items(), key=lambda x: x[1], reverse=True):
            percentage = f"{(count/len(devices)*100):.1f}%"
            type_data.append([dtype.replace("_", " ").title(), str(count), percentage])

        type_table = Table(type_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
        type_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10b981")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f3f4f6")),
                    ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#e5e7eb")),
                ]
            )
        )

        elements.append(type_table)
        elements.append(PageBreak())

        # Network Changes section - highlights differences between scans
        # Only included if change data is provided and contains actual changes
        if changes and any(
            changes.get(k, []) for k in ["new_devices", "missing_devices", "changed_devices"]
        ):
            elements.append(Paragraph("Network Changes", heading_style))

            if changes.get("new_devices"):
                elements.append(
                    Paragraph(f"New Devices ({len(changes['new_devices'])})", styles["Heading3"])
                )
                new_data = [["IP Address", "Hostname", "Type", "Vendor"]]
                for device in changes["new_devices"][:10]:  # Limit to 10 for space
                    new_data.append(
                        [
                            device.get("ip", ""),
                            device.get("hostname", "N/A"),
                            device.get("type", "unknown"),
                            device.get("vendor", "N/A"),
                        ]
                    )

                new_table = Table(
                    new_data, colWidths=[1.5 * inch, 2 * inch, 1.5 * inch, 1.5 * inch]
                )
                new_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10b981")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 10),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f3f4f6")),
                            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#e5e7eb")),
                        ]
                    )
                )
                elements.append(new_table)
                elements.append(Spacer(1, 0.2 * inch))

            if changes.get("missing_devices"):
                elements.append(
                    Paragraph(
                        f"Missing Devices ({len(changes['missing_devices'])})", styles["Heading3"]
                    )
                )
                # Similar table for missing devices

            elements.append(PageBreak())

        # Critical Devices section - emphasizes high-priority assets
        # These devices require special attention for security/availability
        critical_devices = [d for d in devices if d.get("critical", False)]
        if critical_devices:
            elements.append(Paragraph("Critical Infrastructure", heading_style))

            critical_data = [["IP Address", "Hostname", "Type", "Services", "Notes"]]
            for device in critical_devices:
                services = ", ".join(device.get("services", [])[:3])
                if len(device.get("services", [])) > 3:
                    services += "..."

                critical_data.append(
                    [
                        device.get("ip", ""),
                        device.get("hostname", "N/A"),
                        device.get("type", "unknown"),
                        services or "N/A",
                        device.get("notes", "")[:30] + "..."
                        if len(device.get("notes", "")) > 30
                        else device.get("notes", ""),
                    ]
                )

            critical_table = Table(
                critical_data,
                colWidths=[1.2 * inch, 1.5 * inch, 1.2 * inch, 1.8 * inch, 1.8 * inch],
            )
            critical_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ef4444")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#fef2f2")),
                        ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#fecaca")),
                    ]
                )
            )

            elements.append(critical_table)
            elements.append(PageBreak())

        # Detailed Device List - comprehensive inventory table
        # Limited to 50 devices to keep PDF size reasonable
        # Full list available in Excel/CSV exports
        elements.append(Paragraph("Device Inventory", heading_style))
        elements.append(
            Paragraph(
                "Showing first 50 devices. See Excel export for complete list.", styles["Italic"]
            )
        )
        elements.append(Spacer(1, 0.1 * inch))

        device_data = [["IP", "Hostname", "MAC", "Type", "Vendor", "Ports"]]
        for device in devices[:50]:
            ports = ", ".join(map(str, device.get("open_ports", [])[:5]))
            if len(device.get("open_ports", [])) > 5:
                ports += "..."

            device_data.append(
                [
                    device.get("ip", ""),
                    device.get("hostname", "N/A")[:20],
                    device.get("mac", "N/A")[:17],
                    device.get("type", "unknown")[:15],
                    device.get("vendor", "N/A")[:20],
                    ports,
                ]
            )

        device_table = Table(
            device_data,
            colWidths=[1 * inch, 1.5 * inch, 1.3 * inch, 1 * inch, 1.5 * inch, 1.2 * inch],
        )
        device_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366f1")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f3f4f6")),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f9fafb")],
                    ),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )

        elements.append(device_table)

        # Build PDF
        doc.build(elements)

        logger.info(f"PDF report exported to: {filepath}")
        return filepath

    def export_to_excel(
        self, devices: List[Dict], changes: Optional[Dict] = None, filename: Optional[str] = None
    ) -> Path:
        """
        Export network scan data to Excel workbook with multiple sheets.

        Creates a comprehensive Excel workbook with:
        1. Summary sheet - Overview and statistics
        2. All Devices - Complete inventory with formatting
        3. Critical Devices - Filtered view of critical infrastructure
        4. Network Changes - New/missing/changed devices (if data provided)
        5. Subnet Analysis - Network segmentation breakdown

        Excel format provides the most detailed export with:
        - Color coding for device criticality
        - Auto-adjusted column widths
        - Professional formatting and borders
        - Multiple views of the same data
        - Easy filtering and sorting capabilities

        Args:
            devices: List of device dictionaries
            changes: Optional change tracking data
            filename: Optional custom filename

        Returns:
            Path to generated Excel file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filename or f"network_inventory_{timestamp}.xlsx"
        filepath = self.export_path / filename

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Define consistent Excel styles for professional appearance
        # Header style - dark blue background with white text
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Critical device highlighting - light red background
        critical_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        critical_font = Font(color="9C0006")  # Dark red text

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 1. Summary Sheet - Executive overview with key metrics
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet.append(["Network Scan Summary", "", ""])
        summary_sheet.merge_cells("A1:C1")  # Merge for title
        summary_sheet["A1"].font = Font(bold=True, size=16)
        summary_sheet["A1"].alignment = Alignment(horizontal="center")

        summary_sheet.append([])
        summary_sheet.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        summary_sheet.append(["Total Devices:", len(devices)])
        summary_sheet.append(
            ["Critical Devices:", len([d for d in devices if d.get("critical", False)])]
        )

        # Device type breakdown - shows network composition
        summary_sheet.append([])  # Empty row for spacing
        summary_sheet.append(["Device Type", "Count", "Percentage"])

        # Count devices by type for statistical analysis
        device_types = {}
        for device in devices:
            dtype = device.get("type", "unknown")
            device_types[dtype] = device_types.get(dtype, 0) + 1

        row = 8
        for dtype, count in sorted(device_types.items(), key=lambda x: x[1], reverse=True):
            percentage = f"{(count/len(devices)*100):.1f}%"
            summary_sheet.append([dtype.replace("_", " ").title(), count, percentage])
            row += 1

        # Format summary headers
        for cell in summary_sheet["A7:C7"][0]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Adjust column widths
        summary_sheet.column_dimensions["A"].width = 20
        summary_sheet.column_dimensions["B"].width = 15
        summary_sheet.column_dimensions["C"].width = 15

        # 2. All Devices Sheet
        devices_sheet = wb.create_sheet("All Devices")
        headers = [
            "IP Address",
            "Hostname",
            "MAC Address",
            "Vendor",
            "Device Type",
            "Operating System",
            "Critical",
            "Open Ports",
            "Services",
            "Notes",
            "Last Seen",
        ]
        devices_sheet.append(headers)

        # Add device data
        for device in devices:
            ports = ", ".join(map(str, device.get("open_ports", [])))
            services = ", ".join(device.get("services", []))

            row = [
                device.get("ip", ""),
                device.get("hostname", "N/A"),
                device.get("mac", "N/A"),
                device.get("vendor", "N/A"),
                device.get("type", "unknown").replace("_", " ").title(),
                device.get("os", "N/A"),
                "Yes" if device.get("critical", False) else "No",
                ports or "None",
                services or "None",
                device.get("notes", ""),
                device.get("last_seen", datetime.now().isoformat())[:19],
            ]
            devices_sheet.append(row)

        # Format headers
        for cell in devices_sheet["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Highlight critical devices
        for row_num in range(2, len(devices) + 2):
            if devices_sheet[f"G{row_num}"].value == "Yes":
                for col in range(1, 12):
                    cell = devices_sheet.cell(row=row_num, column=col)
                    cell.fill = critical_fill
                    cell.font = critical_font

        # Auto-adjust column widths for readability
        # Calculate optimal width based on content length
        for column in devices_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass  # Skip cells with errors

            # Cap width at 50 to prevent extremely wide columns
            adjusted_width = min(max_length + 2, 50)
            devices_sheet.column_dimensions[column_letter].width = adjusted_width

        # Apply borders to all data cells
        for row in devices_sheet.iter_rows(min_row=1, max_row=len(devices) + 1):
            for cell in row:
                cell.border = border

        # 3. Critical Devices Sheet
        critical_devices = [d for d in devices if d.get("critical", False)]
        if critical_devices:
            critical_sheet = wb.create_sheet("Critical Devices")
            critical_sheet.append(headers)

            for device in critical_devices:
                ports = ", ".join(map(str, device.get("open_ports", [])))
                services = ", ".join(device.get("services", []))

                row = [
                    device.get("ip", ""),
                    device.get("hostname", "N/A"),
                    device.get("mac", "N/A"),
                    device.get("vendor", "N/A"),
                    device.get("type", "unknown").replace("_", " ").title(),
                    device.get("os", "N/A"),
                    "Yes",
                    ports or "None",
                    services or "None",
                    device.get("notes", ""),
                    device.get("last_seen", datetime.now().isoformat())[:19],
                ]
                critical_sheet.append(row)

            # Format critical sheet
            for cell in critical_sheet["1:1"]:
                cell.font = header_font
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.alignment = header_alignment
                cell.border = border

            # Auto-adjust columns
            for column in critical_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                critical_sheet.column_dimensions[column_letter].width = adjusted_width

        # 4. Changes Sheet (if available)
        if changes and any(
            changes.get(k, []) for k in ["new_devices", "missing_devices", "changed_devices"]
        ):
            changes_sheet = wb.create_sheet("Network Changes")

            row_num = 1

            # New devices
            if changes.get("new_devices"):
                changes_sheet.append([f"NEW DEVICES ({len(changes['new_devices'])})", "", "", ""])
                changes_sheet.merge_cells(f"A{row_num}:D{row_num}")
                changes_sheet[f"A{row_num}"].font = Font(bold=True, size=14, color="00B050")
                row_num += 2

                changes_sheet.append(["IP Address", "Hostname", "Type", "Vendor"])
                row_num += 1

                for device in changes["new_devices"]:
                    changes_sheet.append(
                        [
                            device.get("ip", ""),
                            device.get("hostname", "N/A"),
                            device.get("type", "unknown"),
                            device.get("vendor", "N/A"),
                        ]
                    )
                    row_num += 1

                row_num += 2

            # Missing devices
            if changes.get("missing_devices"):
                changes_sheet.append(
                    [f"MISSING DEVICES ({len(changes['missing_devices'])})", "", "", ""]
                )
                changes_sheet.merge_cells(f"A{row_num}:D{row_num}")
                changes_sheet[f"A{row_num}"].font = Font(bold=True, size=14, color="FF0000")
                row_num += 2

                changes_sheet.append(["IP Address", "Hostname", "Type", "Last Seen"])
                row_num += 1

                for device in changes["missing_devices"]:
                    changes_sheet.append(
                        [
                            device.get("ip", ""),
                            device.get("hostname", "N/A"),
                            device.get("type", "unknown"),
                            device.get("last_seen", "N/A"),
                        ]
                    )
                    row_num += 1

        # 5. Subnet Analysis Sheet - Network segmentation view
        # Helps identify network organization and potential issues
        subnet_sheet = wb.create_sheet("Subnet Analysis")
        subnet_data = self._analyze_subnets(devices)

        # Headers show device distribution by subnet
        subnet_sheet.append(
            ["Subnet", "Device Count", "Router", "Switches", "Servers", "Workstations", "Other"]
        )

        for subnet, data in sorted(subnet_data.items()):
            subnet_sheet.append(
                [
                    subnet,
                    data["total"],
                    data["types"].get("router", 0),
                    data["types"].get("switch", 0),
                    data["servers"],
                    data["types"].get("workstation", 0),
                    data["other"],
                ]
            )

        # Format subnet sheet
        for cell in subnet_sheet["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Save workbook
        wb.save(filepath)

        logger.info(f"Excel report exported to: {filepath}")
        return filepath

    def export_to_json(
        self, devices: List[Dict], changes: Optional[Dict] = None, filename: Optional[str] = None
    ) -> Path:
        """
        Export to enhanced JSON format with metadata.

        JSON export provides:
        - Complete data preservation
        - Machine-readable format
        - Metadata for context
        - Structured organization
        - Easy integration with other tools

        The export includes summary statistics, subnet analysis,
        and all device data in a hierarchical structure suitable
        for APIs and data processing pipelines.

        Args:
            devices: List of device dictionaries
            changes: Optional change tracking data
            filename: Optional custom filename

        Returns:
            Path to generated JSON file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filename or f"network_export_{timestamp}.json"
        filepath = self.export_path / filename

        # Prepare comprehensive export data structure
        export_data = {
            "metadata": {
                "export_date": datetime.now().isoformat(),
                "total_devices": len(devices),
                "critical_devices": len([d for d in devices if d.get("critical", False)]),
                "exporter_version": "2.0",
            },
            "summary": self._generate_summary(devices),  # Statistical overview
            "devices": devices,  # Complete device inventory
            "changes": changes or {},  # Network changes if available
            "subnet_analysis": self._analyze_subnets(devices),  # Network segmentation
        }

        with open(filepath, "w") as f:
            json.dump(export_data, f, indent=2, default=str)

        logger.info(f"JSON export saved to: {filepath}")
        return filepath

    def export_to_csv_enhanced(self, devices: List[Dict], filename: Optional[str] = None) -> Path:
        """
        Export to enhanced CSV format with vulnerability data.

        CSV export is ideal for:
        - Importing into spreadsheet applications
        - Data analysis tools
        - Simple interchange format
        - Wide compatibility

        Includes vulnerability information and flattens nested
        data structures for tabular representation.

        Args:
            devices: List of device dictionaries
            filename: Optional custom filename

        Returns:
            Path to generated CSV file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = filename or f"devices_{timestamp}.csv"
        filepath = self.export_path / filename

        # Convert to DataFrame for better handling
        df = pd.DataFrame(devices)

        # Flatten nested fields
        if "open_ports" in df.columns:
            df["open_ports"] = df["open_ports"].apply(
                lambda x: ";".join(map(str, x)) if isinstance(x, list) else ""
            )

        if "services" in df.columns:
            df["services"] = df["services"].apply(
                lambda x: ";".join(x) if isinstance(x, list) else ""
            )

        # Extract data from API intelligence
        if "api_intelligence" in df.columns:
            # Extract data sources
            df["vendor_source"] = df["api_intelligence"].apply(
                lambda x: x.get("data_sources", {}).get("mac_vendor", "")
                if isinstance(x, dict)
                else ""
            )

            # Extract API confidence
            df["api_confidence"] = df["api_intelligence"].apply(
                lambda x: x.get("confidence", 0) if isinstance(x, dict) else 0
            )

            # Drop the complex api_intelligence column
            df = df.drop(columns=["api_intelligence"])

        # Extract passive analysis data if present
        if "passive_analysis" in df.columns:
            df["traffic_flows"] = df["passive_analysis"].apply(
                lambda x: x.get("traffic_flows", 0) if isinstance(x, dict) else 0
            )
            df["passive_services"] = df["passive_analysis"].apply(
                lambda x: ";".join(x.get("services_observed", [])) if isinstance(x, dict) else ""
            )
            df = df.drop(columns=["passive_analysis"])

        # Save to CSV
        df.to_csv(filepath, index=False)

        logger.info(f"Enhanced CSV exported to: {filepath}")
        return filepath

    def _generate_summary(self, devices: List[Dict]) -> Dict:
        """
        Generate comprehensive summary statistics from device data.

        Analyzes the device list to produce insights including:
        - Device type distribution
        - Vendor breakdown
        - Operating system diversity
        - Common services and ports
        - Security risk indicators (high-risk ports)

        This summary helps identify:
        - Network composition patterns
        - Potential security concerns
        - Vendor consolidation opportunities
        - Service standardization needs

        Args:
            devices: List of device dictionaries

        Returns:
            Summary statistics dictionary with categorized metrics
        """
        summary = {
            "total_devices": len(devices),
            "device_types": {},
            "critical_devices": 0,
            "vendors": {},
            "operating_systems": {},
            "common_services": {},
            "port_statistics": {"most_common": {}, "high_risk_ports": 0},
        }

        # Define high-risk ports that may indicate security concerns
        # These ports are commonly targeted or expose sensitive services
        high_risk_ports = {21, 23, 135, 139, 445, 3389, 5900}  # FTP, Telnet, NetBIOS, RDP, VNC
        all_ports = []  # Collect all ports for frequency analysis

        for device in devices:
            # Device types
            dtype = device.get("type", "unknown")
            summary["device_types"][dtype] = summary["device_types"].get(dtype, 0) + 1

            # Critical devices
            if device.get("critical", False):
                summary["critical_devices"] += 1

            # Vendors
            vendor = device.get("vendor", "Unknown")
            if vendor:
                summary["vendors"][vendor] = summary["vendors"].get(vendor, 0) + 1

            # Operating systems
            os = device.get("os", "Unknown")
            if os:
                summary["operating_systems"][os] = summary["operating_systems"].get(os, 0) + 1

            # Services
            for service in device.get("services", []):
                service_name = service.split(":")[0] if ":" in service else service
                summary["common_services"][service_name] = (
                    summary["common_services"].get(service_name, 0) + 1
                )

            # Ports
            ports = device.get("open_ports", [])
            all_ports.extend(ports)

            # Check for high-risk ports
            if any(port in high_risk_ports for port in ports):
                summary["port_statistics"]["high_risk_ports"] += 1

        # Calculate most common ports
        from collections import Counter

        port_counts = Counter(all_ports)
        summary["port_statistics"]["most_common"] = dict(port_counts.most_common(10))

        return summary

    def _analyze_subnets(self, devices: List[Dict]) -> Dict:
        """
        Analyze device distribution across network subnets.

        Groups devices by /24 subnets to understand:
        - Network segmentation
        - Device density by subnet
        - Critical infrastructure distribution
        - Subnet-specific device types

        This analysis helps with:
        - Network architecture documentation
        - Identifying subnet purposes
        - Security zone planning
        - Capacity planning

        Note: Assumes /24 subnet masks for simplicity. Real subnet
        masks may vary in production networks.

        Args:
            devices: List of device dictionaries

        Returns:
            Dictionary mapping subnets to device statistics
        """
        subnets = {}

        for device in devices:
            ip = device.get("ip", "")
            if not ip:
                continue

            # Extract subnet (assuming /24 for standardization)
            # This simplification works for most common network designs
            parts = ip.split(".")
            if len(parts) == 4:
                subnet = f"{parts[0]}.{parts[1]}.{parts[2]}.0/24"

                if subnet not in subnets:
                    subnets[subnet] = {
                        "total": 0,
                        "types": {},
                        "critical": 0,
                        "servers": 0,
                        "other": 0,
                    }

                subnets[subnet]["total"] += 1

                # Device type
                dtype = device.get("type", "unknown")
                subnets[subnet]["types"][dtype] = subnets[subnet]["types"].get(dtype, 0) + 1

                # Critical devices
                if device.get("critical", False):
                    subnets[subnet]["critical"] += 1

                # Categorize devices for subnet role analysis
                # Helps identify subnet purposes (server farm, user segment, etc.)
                if "server" in dtype:
                    subnets[subnet]["servers"] += 1
                elif dtype not in ["router", "switch", "workstation"]:
                    subnets[subnet]["other"] += 1  # IoT, printers, other devices

        return subnets
